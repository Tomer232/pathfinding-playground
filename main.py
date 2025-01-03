import pygame
import numpy as np
import random
import os
import heapq
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from collections import deque

class Grid:
    def __init__(self, screen_width=800, screen_height=600, cell_size=7):
        self.screen_width = screen_width
        self.screen_height = screen_height
        self.cell_size = cell_size

        self.width = screen_width // cell_size
        self.height = screen_height // cell_size

        self.grid = np.zeros((self.height, self.width), dtype=int)

        self._create_outer_walls()

        self._generate_maze()

        # Colors
        self.COLORS = {
            0: (255, 255, 255),  # Empty cell
            1: (0, 0, 0),  # Wall
            2: (0, 255, 0),  # Start point
            3: (255, 0, 0),  # End point
            4: (114, 9, 183),  # Path
            5: (236, 188, 253)  # Visited cell
        }

    def _create_outer_walls(self):
        self.grid[0, :] = 1
        self.grid[-1, :] = 1

        self.grid[:, 0] = 1
        self.grid[:, -1] = 1

    def _generate_maze(self):
        for y in range(1, self.height - 1):
            for x in range(1, self.width - 1):
                if random.random() < 0.3:  # 30% chance to place a wall
                    self.grid[y, x] = 1

    def place_start_and_end_points(self):
        self.start_point = self._get_random_safe_position()
        self.end_point = self._get_random_safe_position(far_from=self.start_point, min_distance=30)

        self.grid[self.start_point[1], self.start_point[0]] = 2  # Start point
        self.grid[self.end_point[1], self.end_point[0]] = 3      # End point

    def _get_random_safe_position(self, far_from=None, min_distance=0):
        while True:
            x = random.randint(1, self.width - 2)
            y = random.randint(1, self.height - 2)

            if self.grid[y, x] == 0:
                if far_from:
                    distance = abs(far_from[0] - x) + abs(far_from[1] - y)
                    if distance >= min_distance:
                        return x, y
                else:
                    return x, y

    def draw(self, screen):
        for y in range(self.height):
            for x in range(self.width):
                rect = pygame.Rect(
                    x * self.cell_size,
                    y * self.cell_size,
                    self.cell_size,
                    self.cell_size
                )
                pygame.draw.rect(screen, self.COLORS[self.grid[y][x]], rect)
                pygame.draw.rect(screen, (200, 200, 200), rect, 1)

    def reset_visualization(self):
        for y in range(self.height):
            for x in range(self.width):
                if self.grid[y, x] in [4, 5]:  
                    self.grid[y, x] = 0

        self.grid[self.start_point[1], self.start_point[0]] = 2
        self.grid[self.end_point[1], self.end_point[0]] = 3

    def bfs(self, screen):
        start = self.start_point
        end = self.end_point
        
        directions = [(0, 1), (1, 0), (0, -1), (-1, 0),
                    (1, 1), (-1, -1), (1, -1), (-1, 1)]
        
        visited = set([start])  
        queue = deque([(start, [start])])  
        update_frequency = 10  
        updates = 0
        
        while queue:
            current, path = queue.popleft()
            x, y = current
            
            if current == end:
                for px, py in path:
                    if self.grid[py][px] not in [2, 3]:
                        self.grid[py][px] = 4
                return True
            
            if self.grid[y][x] not in [2, 3]:
                self.grid[y][x] = 5
                updates += 1
                
                if updates % update_frequency == 0:
                    self.draw(screen)
                    pygame.display.flip()
            
            for dx, dy in directions:
                nx, ny = x + dx, y + dy
                neighbor = (nx, ny)
                
                if (0 <= nx < self.width and 0 <= ny < self.height and neighbor not in visited and self.grid[ny][nx] != 1):
                    visited.add(neighbor)
                    new_path = path + [neighbor]
                    queue.append((neighbor, new_path))
                    
        return False

    def dfs(self, screen):
        start = self.start_point
        end = self.end_point
        
        directions = [(0, 1), (1, 0), (0, -1), (-1, 0),
                    (1, 1), (-1, -1), (1, -1), (-1, 1)]
        random.shuffle(directions)  
        
        visited = set([start])  
        stack = [(start, [start])] 
        update_frequency = 10  
        updates = 0
        
        while stack:
            current, path = stack.pop()
            x, y = current
            
            if current == end:
                for px, py in path:
                    if self.grid[py][px] not in [2, 3]:
                        self.grid[py][px] = 4
                return True
            
            if self.grid[y][x] not in [2, 3]:
                self.grid[y][x] = 5
                updates += 1
                
                if updates % update_frequency == 0:
                    self.draw(screen)
                    pygame.display.flip()
                    pygame.event.pump()  
            
            if len(path) % 5 == 0:
                random.shuffle(directions)
            
            valid_moves = []
            for dx, dy in directions:
                nx, ny = x + dx, y + dy
                neighbor = (nx, ny)
                
                if (0 <= nx < self.width and 
                    0 <= ny < self.height and 
                    neighbor not in visited and 
                    self.grid[ny][nx] != 1):
                    
                    valid_moves.append(neighbor)
            
            for move in valid_moves:
                visited.add(move)
                stack.append((move, path + [move]))
        
        return False

    def a_star(self, screen):
        start = self.start_point
        end = self.end_point
        
        def heuristic(current, goal):
            # Manhattan distance heuristic
            return abs(current[0] - goal[0]) + abs(current[1] - goal[1])

        directions = [(0, 1), (1, 0), (0, -1), (-1, 0), 
                    (1, 1), (-1, -1), (1, -1), (-1, 1)]
        
        g_scores = {start: 0}
        f_scores = {start: heuristic(start, end)}
        came_from = {}
        
        open_set = [(f_scores[start], start)]
        closed_set = set()

        while open_set:
            current = heapq.heappop(open_set)[1]
            
            if current == end:
                path = []
                while current in came_from:
                    if self.grid[current[1]][current[0]] not in [2, 3]:
                        self.grid[current[1]][current[0]] = 4
                    current = came_from[current]
                return True
                
            closed_set.add(current)
            
            for dx, dy in directions:
                neighbor = (current[0] + dx, current[1] + dy)
                
                if (not (0 <= neighbor[0] < self.width and 
                        0 <= neighbor[1] < self.height) or
                    neighbor in closed_set or
                    self.grid[neighbor[1]][neighbor[0]] == 1):
                    continue

                tentative_g = g_scores[current] + (1.414 if dx and dy else 1)

                if (neighbor not in g_scores or 
                    tentative_g < g_scores[neighbor]):
                    came_from[neighbor] = current
                    g_scores[neighbor] = tentative_g
                    f_scores[neighbor] = tentative_g + heuristic(neighbor, end)
                    heapq.heappush(open_set, (f_scores[neighbor], neighbor))
                    
                    if self.grid[neighbor[1]][neighbor[0]] not in [2, 3]:
                        self.grid[neighbor[1]][neighbor[0]] = 5
                        
                    if len(closed_set) % 10 == 0:
                        self.draw(screen)
                        pygame.display.flip()

        return False

    def ib_rrt_star(self, screen):
        start = self.start_point
        end = self.end_point
        
        class Node:
            def __init__(self, position, parent=None, cost=0.0):
                self.position = position
                self.parent = parent
                self.cost = cost
        
        def distance(pos1, pos2):
            return ((pos1[0] - pos2[0]) ** 2 + (pos1[1] - pos2[1]) ** 2) ** 0.5
        
        def is_valid(x, y):
            return (0 <= x < self.width and 
                    0 <= y < self.height and 
                    self.grid[y][x] != 1)
        
        def interpolate_path(start_pos, end_pos):
            path = []
            x1, y1 = start_pos
            x2, y2 = end_pos
            dx = abs(x2 - x1)
            dy = abs(y2 - y1)
            x, y = x1, y1
            
            step_x = 1 if x1 < x2 else -1
            step_y = 1 if y1 < y2 else -1
            
            if dx > dy:
                err = dx / 2.0
                while x != x2:
                    path.append((x, y))
                    err -= dy
                    if err < 0:
                        y += step_y
                        err += dx
                    x += step_x
            else:
                err = dy / 2.0
                while y != y2:
                    path.append((x, y))
                    err -= dx
                    if err < 0:
                        x += step_x
                        err += dy
                    y += step_y
            
            path.append((x2, y2))
            return path
        
        def check_path(path):
            for x, y in path:
                if not is_valid(x, y):
                    return False
            return True
        
        def find_nearest_node(tree, position):
            return min(tree.values(), key=lambda node: distance(node.position, position))
        
        def extend_tree(tree, target_pos):
            nearest_node = find_nearest_node(tree, target_pos)
            current_pos = nearest_node.position
            
            dx = target_pos[0] - current_pos[0]
            dy = target_pos[1] - current_pos[1]
            dist = (dx * dx + dy * dy) ** 0.5
            
            if dist == 0:
                return None
            
            step_size = 3
            dx = int(round(dx / dist * step_size))
            dy = int(round(dy / dist * step_size))
            
            new_x = current_pos[0] + dx
            new_y = current_pos[1] + dy
            new_pos = (new_x, new_y)
            
            if not is_valid(new_x, new_y):
                return None
            
            path = interpolate_path(current_pos, new_pos)
            if not check_path(path):
                return None
            
            new_cost = nearest_node.cost + distance(current_pos, new_pos)
            new_node = Node(new_pos, nearest_node, new_cost)
            tree[new_pos] = new_node
            
            for px, py in path:
                if self.grid[py][px] not in [2, 3]:
                    self.grid[py][px] = 5
            
            return new_node
        
        def construct_path(node):
            path = []
            current = node
            while current is not None:
                path.append(current.position)
                current = current.parent
            return path[::-1]
        
        def visualize_final_path(path_points):
            for i in range(len(path_points) - 1):
                point1 = path_points[i]
                point2 = path_points[i + 1]
                
                interpolated = interpolate_path(point1, point2)
                for x, y in interpolated:
                    if self.grid[y][x] not in [2, 3]:  
                        self.grid[y][x] = 4
        
        start_tree = {start: Node(start)}
        end_tree = {end: Node(end)}
        
        max_iterations = 2000
        goal_bias = 0.2
        
        for iteration in range(max_iterations):
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    return False
            
            if iteration % 2 == 0:
                active_tree = start_tree
                other_tree = end_tree
            else:
                active_tree = end_tree
                other_tree = start_tree
            
            if random.random() < goal_bias:
                target = list(other_tree.keys())[0]
            else:
                target = (random.randint(1, self.width-2),
                        random.randint(1, self.height-2))
            
            new_node = extend_tree(active_tree, target)
            
            if new_node:
                for other_pos, other_node in other_tree.items():
                    if distance(new_node.position, other_pos) < 5:
                        connecting_path = interpolate_path(new_node.position, other_pos)
                        if check_path(connecting_path):
                            # Construct complete path
                            if active_tree == start_tree:
                                path_points = (construct_path(new_node) + 
                                            connecting_path +
                                            construct_path(other_node)[::-1])
                            else:
                                path_points = (construct_path(other_node) + 
                                            connecting_path +
                                            construct_path(new_node)[::-1])
                            
                            visualize_final_path(path_points)
                            return True
            
            if iteration % 10 == 0:
                self.draw(screen)
                pygame.display.flip()
        
        return False




def save_playground_with_data(screen, playground_name, best_algorithm):
    base_path = r"C:\Users\tomer\Desktop\Projects\pathfinding"
    excel_file = os.path.join(base_path, "playground_metadata.xlsx")
    image_folder = os.path.join(base_path, "playground_images")

    if not os.path.exists(image_folder):
        os.makedirs(image_folder)

    image_path = os.path.join(image_folder, f"{playground_name}.png")
    pygame.image.save(screen, image_path)

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Playground Name", "Date", "Time", "Best Algorithm", "Image"])

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    next_row = ws.max_row + 7
    ws.cell(row=next_row, column=1, value=playground_name)
    ws.cell(row=next_row, column=2, value=date)
    ws.cell(row=next_row, column=3, value=time)
    ws.cell(row=next_row, column=4, value=best_algorithm)

    img = ExcelImage(image_path)
    img.anchor = f"E{next_row}"
    img.width = 160
    img.height = 120
    ws.add_image(img)

    wb.save(excel_file)
    print(f"Playground and data saved in '{excel_file}'")


current_algorithm_running = False


def main():
    global current_algorithm_running

    pygame.init()

    SCREEN_WIDTH = 1200
    SCREEN_HEIGHT = 600

    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
    pygame.display.set_caption("Pathfinding Playground")

    grid = Grid(cell_size=7)
    grid.place_start_and_end_points()

    buttons = [
        {"name": "BFS", "rect": pygame.Rect(900, 50, 200, 50)},
        {"name": "DFS", "rect": pygame.Rect(900, 120, 200, 50)},
        {"name": "A*", "rect": pygame.Rect(900, 190, 200, 50)},
        {"name": "IB-RRT*", "rect": pygame.Rect(900, 260, 200, 50)},
        {"name": "NEW MAP", "rect": pygame.Rect(900, 330, 200, 50)}
    ]


    def draw_buttons(running_algo = None):
        font = pygame.font.SysFont("Verdana", 36)
        button_color = (232, 241, 242)
        new_map_color = (34, 177, 76)
        outline_color = (0, 0, 0)
        text_color = (0, 0, 0)
        active_button_color = (4, 150, 255)
        active_text_color = (255, 255, 255)
        panel_color = (108, 117, 107)

        panel_rect = pygame.Rect(798, 0, 450, SCREEN_HEIGHT)
        pygame.draw.rect(screen, panel_color, panel_rect)

        for button in buttons:
            if running_algo == button["name"]:  # Highlight the currently running button
                pygame.draw.rect(screen, outline_color, button["rect"], 2)

                if button["name"] == "NEW MAP":
                    pygame.draw.rect(screen, new_map_color, button["rect"].inflate(-4, -4))
                else:
                    pygame.draw.rect(screen, active_button_color, button["rect"].inflate(-4, -4))
                text = font.render(button["name"], True, active_text_color)
            else:
                pygame.draw.rect(screen, outline_color, button["rect"], 2)
                pygame.draw.rect(screen, button_color, button["rect"].inflate(-4, -4))
                text = font.render(button["name"], True, text_color)

            text_rect = text.get_rect(center=button["rect"].center)
            screen.blit(text, text_rect)

    screen.fill((255, 255, 255))
    grid.draw(screen)
    draw_buttons()

    running = True
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            elif event.type == pygame.MOUSEBUTTONDOWN:
                mouse_pos = event.pos
                for button in buttons:
                    if button["rect"].collidepoint(mouse_pos):
                        if button["name"] == "BFS" and not current_algorithm_running:
                            current_algorithm_running = True
                            grid.reset_visualization()
                            draw_buttons(running_algo="BFS")  
                            pygame.display.flip()  
                            if grid.bfs(screen):
                                save_playground_with_data(screen, "playground_bfs", "BFS")
                            current_algorithm_running = False
                        elif button["name"] == "DFS" and not current_algorithm_running:
                            current_algorithm_running = True
                            grid.reset_visualization()
                            draw_buttons(running_algo="DFS")  
                            pygame.display.flip() 
                            if grid.dfs(screen):
                                save_playground_with_data(screen, "playground_dfs", "DFS")
                            current_algorithm_running = False
                        elif button["name"] == "A*" and not current_algorithm_running:
                            current_algorithm_running = True
                            grid.reset_visualization()
                            draw_buttons(running_algo="A*")  
                            pygame.display.flip()  
                            if grid.a_star(screen):
                                save_playground_with_data(screen, "playground_astar", "A*")
                            current_algorithm_running = False
                        elif button["name"] == "IB-RRT*" and not current_algorithm_running:
                            current_algorithm_running = True
                            grid.reset_visualization()
                            draw_buttons(running_algo="IB-RRT*")  
                            pygame.display.flip()  
                            if grid.ib_rrt_star(screen):
                                save_playground_with_data(screen, "playground_ib_rrt_star", "IB-RRT*")
                            current_algorithm_running = False
                        elif button["name"] == "NEW MAP":
                            current_algorithm_running == True
                            draw_buttons(running_algo="NEW MAP")
                            pygame.display.flip()
                            grid = Grid(cell_size=7)
                            grid.place_start_and_end_points()
                            current_algorithm_running = False

        screen.fill((255, 255, 255))
        grid.draw(screen)
        draw_buttons()
        pygame.display.flip()

    pygame.quit()


if __name__ == "__main__":
    main()